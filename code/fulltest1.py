import mariadb
import sys
import openpyxl
import getpass

from tkinter import *
from tkinter import ttk, filedialog
from tkinter.filedialog import askopenfile
import os


# Establish connection to database
try:
	conn = mariadb.connect(
		user="topplab",
		password=getpass.getpass(prompt='Database user password: '),
		host="10.16.0.101",
		port=3306,
		database="tag_server"
	)
except mariadb.Error as e:
    print(f"Error connecting to MariaDB Platform: {e}")
    sys.exit(1)

cur = conn.cursor()

path = sys.argv[1]
wb = openpyxl.load_workbook(path, data_only=True)

sheet = wb.active 

row = sheet.max_row #rows in sheet
column = sheet.max_column #cols in sheet


def insertValue(dbTable,dbCol,value):
    '''
    Inserts a single value to the connected database

    Parameters:
        str dbTable (name of table in connected database)
        str dbCol (name of col in connected database)
        str value (what is to be inserted)

    Return:
        void

    '''
    assert conn != None, "No database connection"

	query = f"INSERT INTO {dbTable} ({dbCol}) VALUES (?)"

	val = (value,)

	cur.execute(query,val)

	
	conn.commit()



def writeToCell(value,col,row):
	cell = sheet.cell(row=row, column=col)
	cell.value = value

def propagateUid(firstId,lastId,idCol):
	for i in range(firstId,lastId):
		writeToCell(i,idCol,i+1)



def insertAllCellsInCol(colNum):
    '''
    Reads an entire column from open xlsx sheet and inserts all values in sequential order to connected database.

    Parameters:
        int colNum (The column of the database intended to be uploaded. Column numbers increment from 1 starting on left of xlsx document. (A is 1, B is 2, etc.))

    Return:
        void
    '''

	confirm = input("Are you sure you want to insert " + str(row) + " values to the database?\nType YES to continue.\n")
	if confirm != "YES":
		print("Exiting...")
		sys.exit(1)

	for i in range(2, row + 1):
		cell = sheet.cell(row = i, column = colNum)
		insertValue("tags","UUid",cell.value)

	lastId = cur.lastrowid
	firstNewId = lastId - row + 2

	print(lastId)
	print(firstNewId)

	propagateUid(firstNewId,lastId,12)

	# wb.save(path)

	conn.commit()

	
#Deprecated
# def getLastId():
# 	query = f"SELECT id FROM tags WHERE id=(SELECT max(id) FROM tags)"
# 	value = ()
# 	cur.execute(query,value)
# 	return 
	



insertAllCellsInCol(11)
# print(getLastId())
#insertValue("tags","UUid","test393966")
cur.close()

conn.close()