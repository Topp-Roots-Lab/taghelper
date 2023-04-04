#-Dylan Fritz <dfritz1211@gmail.com>
import mariadb
import sys
import openpyxl
from openpyxl import Workbook

from tkinter import *
from tkinter import ttk, filedialog
from tkinter.filedialog import askopenfile
import os

import getpass

LASTID = None
FIRSTNEWID = None
FILE_PATH = None

try:
    conn = mariadb.connect(
        user=os.getlogin(),
        password=getpass.getpass(prompt='Database user password: '),
        host="10.16.0.101", #Nebula's relational ip!
        port=3306,
        database="tag_server"
    )
except mariadb.Error as e:
    print(f"Error connecting to MariaDB Platform: {e}")
    sys.exit(1)

cur = conn.cursor()



# Create an instance of tkinter frame
win = Tk()

# Set the geometry of tkinter frame based on screen size
win.title("TagHelper")
win.configure(bg='#b3d98d')

normal_width = 1920 
normal_height = 1080
screen_width = win.winfo_screenwidth()
screen_height = win.winfo_screenheight()
percent_width = screen_width / (normal_width / 100)
percent_height = screen_height / (normal_height / 100)

scale_factor = ((percent_width + percent_height) / 2) / 100
fontsize = int(14*scale_factor)
minimum_size = 8
if fontsize < minimum_size:
    fontsize = minimum_size

buttonStyle = ttk.Style()
buttonStyle.configure('TButton', font=("Helvetica", fontsize, ))

frameStyle = ttk.Style()
frameStyle.configure('TFrame', background="Teal")

frame = ttk.Frame(win, padding=int(10*scale_factor), style='TFrame')

frame.grid(row=0, column=0, sticky="")

def insertValue(dbTable,dbCol,value):
    '''
    Inserts a single value to the connected database

    Parameters:
        str dbTable (name of table in connected database)
        str dbCol (name of col in connected database)
        str value (what is to be inserted)

    Return:
        void

    '''
    assert conn != None, "No database connection"

    query = f"INSERT INTO {dbTable} ({dbCol}) VALUES (?)"

    val = (value,)

    cur.execute(query,val)

    
    conn.commit()

def open_file():
    '''
    Opens a choose file widget
    '''
    global FILE_PATH
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel', '*.xlsx')])
    if file:
      filepath = os.path.relpath(file.name)
      print(filepath)
      FILE_PATH = filepath
 
      assert FILE_PATH != None, "Should have a file path"
      assert os.path.isfile(FILE_PATH)
      
      Label(frame, text=str(filepath), font=('Aerial 11')).pack()

def insertAllCellsInCol(path, colNum, sheetName, firstDataRow=1):
    global FIRSTNEWID
    global LASTID
    '''
    Reads an entire column from open xlsx sheet and inserts all values in sequential order to connected database.

    Parameters:
        str path (The path to the .xlsx finle to upload)
        int colNum (The column of the database intended to be uploaded. Column numbers increment from 1 starting on left of xlsx document. (A is 1, B is 2, etc.))
        str sheetName (The name of the excel sheet to be witten to.)
        int firstDataRow (The first row in the spreadsheet containing data needing to be pulled. Default value 1. Overridden in examples where a spreadsheet contains headings on the first few rows etc.)


    Return:
        void
    '''
    
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheetName]

    row = int(last_row_num_E_bc.get()) #rows in sheet
    column = ws.max_column #cols in sheet

    print("MAX ROWS: "+str(row))


    # confirm = input("Are you sure you want to insert " + str(row) + " values to the database?\nType YES to continue.\n")
    # if confirm != "YES":
    #     print("Exiting...")


    for i in range(firstDataRow, row+frame) : 

        print("READING ROW: " + str(i))
        
        cell = ws.cell(row = i, column = colNum)
        insertValue("tags","UUid",cell.value)

    lastId = cur.lastrowid
    firstNewId = lastId - row + firstDataRow

    # print(lastId)
    # print(firstNewId)

    FIRSTNEWID =firstNewId
    LASTID = lastId


    # wb.save(path)

    conn.commit()
    wb.close()




def write_wb(path, firstid, lastid, col_n, sheetName, startrow=1):
    '''
    Writes to a column (col_n) of (path).xlsx the values from (firstid) to (lastid) starting on row (startrow) and incrementing down the column.

    Parameters:
        str path (The path to the .xlsx file to write to. Must end in .xlsx.)
        int firstid (The first uid assigned to the first row to be written to.)
        int lastid (The last id belonging to the last row in this sheet.)
        int col_n (The column to write id values to. Sheet columns are idexed from left to right starting at 1. (A is 1, B is 2, etc.))
        str sheetName (The name of the excel sheet to be witten to.)
        int startrow [optional] (The first row that contains data needing an id value (not col headers etc.). Default value 1 (row 1).)

    Return:
        void

    '''
    wb = openpyxl.load_workbook(path, data_only=True)
    
   
    ws = wb[sheetName]
    rowi = startrow

    assert firstid != None, "Must insert values to generate ids"
    assert lastid != None, "Must insert values to generate ids"

    for i in range(firstid, lastid+1):
        cell = ws.cell(row = rowi, column = col_n)
        cell.value = str(i)
        rowi+=1

    wb.save(path)
    wb.close()
    print("write successful")


#tkinter window stuff
file_label = Label(frame, text="Select a file to upload", font=('Helvetica', fontsize)).grid(row=1, column=0)

ttk.Button(frame, text="Browse", style='TButton', command=open_file).grid(row=2, column=0)

sheetname_label = Label(frame, text="Enter the sheet name containing barcode strings EXACTLY AS IT IS IN EXCEL:", font=('Helvetica', fontsize)).grid(row=3, column=0)

sheet_name_E = Entry(frame,font=('Helvetica', fontsize)).grid(row=4, column=0)

colnum_label = Label(frame, text="Enter the col number containing barcode strings:", font=('Helvetica', fontsize)).grid(row=5, column=0)

col_num_E = Entry(frame,font=('Helvetica', fontsize)).grid(row=6, column=0)

rownum_label_bc = Label(frame, text="Enter the first row number containing barcode strings:", font=('Helvetica', fontsize)).grid(row=7, column=0)

row_num_E_bc = Entry(frame, font=('Helvetica', fontsize)).grid(row=8, column=0)

last_rownum_label_bc = Label(frame, text="Enter the last row number containing barcode strings:", font=('Helvetica', fontsize)).grid(row=9, column=0)

last_row_num_E_bc = Entry(frame,font=('Helvetica', fontsize)).grid(row=10, column=0)

ttk.Button(frame, text="Upload", style='TButton', command=lambda: insertAllCellsInCol(FILE_PATH, int(col_num_E.get()), str(sheet_name_E.get()), firstDataRow=int(row_num_E_bc.get()))).grid(row=11, column=0)

val_label = Label(frame, text="Enter the col number that will contain Uids.", font=('Helvetica', fontsize)).grid(row=12, column=0)

val_E = Entry(frame,font=('Helvetica', fontsize)).grid(row=13, column=0)

rownum_label = Label(frame, text="Enter the first row number requiring a Uid", font=('Helvetica', fontsize)).grid(row=14, column=0)

row_num_E = Entry(frame,font=('Helvetica', fontsize)).grid(row=15, column=0)

ttk.Button(frame, text="Write", style='TButton', command=lambda: write_wb(FILE_PATH, FIRSTNEWID, LASTID, int(val_E.get()), str(sheet_name_E.get()), startrow=int(row_num_E.get()))).grid(row=16, column=0)

for child in frame.winfo_children():
    child.grid_configure(padx=int(10*scale_factor), pady=int(10*scale_factor))

win.mainloop()




