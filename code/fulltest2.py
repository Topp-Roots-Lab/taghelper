import mariadb
import sys
import openpyxl
from openpyxl import Workbook

from tkinter import *
from tkinter import ttk, filedialog
from tkinter.filedialog import askopenfile
import os



try:
	conn = mariadb.connect(
		user="topplab",
		password=getpass.getpass(prompt='Database user password: '),
		host="10.16.0.101",
		port=3306,
		database="tag_server"
	)
except mariadb.Error as e:
    print(f"Error connecting to MariaDB Platform: {e}")
    sys.exit(1)

FILE_PATH = None

# Create an instance of tkinter frame
win = Tk()

# Set the geometry of tkinter frame
win.geometry("700x350")

def open_file():
    global FILE_PATH
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel', '*.xlsx')])
    if file:
      filepath = os.path.relpath(file.name)
      print(filepath)
      FILE_PATH = filepath
 
      assert FILE_PATH != None, "Should have a file path"
      assert os.path.isfile(FILE_PATH)
      
      Label(win, text=str(filepath), font=('Aerial 11')).pack()

def insertAllCellsInCol(path, colNum):
    '''
    Reads an entire column from open xlsx sheet and inserts all values in sequential order to connected database.

    Parameters:
        int colNum (The column of the database intended to be uploaded. Column numbers increment from 1 starting on left of xlsx document. (A is 1, B is 2, etc.))

    Return:
        void
    '''
    global FILE_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active


	confirm = input("Are you sure you want to insert " + str(row) + " values to the database?\nType YES to continue.\n")
	if confirm != "YES":
		print("Exiting...")
		sys.exit(1)

	for i in range(2, row + 1):
		cell = ws.cell(row = i, column = colNum)
		insertValue("tags","UUid",cell.value)

	lastId = cur.lastrowid
	firstNewId = lastId - row + 2

	print(lastId)
	print(firstNewId)


	# wb.save(path)

	conn.commit()




def write_wb(path, row_n, col_n, data):
    wb = openpyxl.load_workbook(path, data_only=True)
    
   
    ws = wb.active
    for i in range(1, row_n):
        cell = ws.cell(row = i, column = col_n)
        cell.value = str(data)

    wb.save(path)
    print("write successful")


# Add a Label widget
file_label = Label(win, text="Select a file to upload", font=('Georgia 13')).pack(pady=10)



# Create a Button
ttk.Button(win, text="Browse", command=open_file).pack(pady=20)

colnum_label = Label(win, text="Enter the col number", font=('Georgia 13')).pack(pady=10)


col_num_E = Entry(win,font=('Georgia 13'),width=40)
col_num_E.pack(pady=20)

rownum_label = Label(win, text="Enter the row number", font=('Georgia 13')).pack(pady=10)

row_num_E = Entry(win,font=('Georgia 13'),width=40)
row_num_E.pack(pady=20)

val_label = Label(win, text="Enter the value to write", font=('Georgia 13')).pack(pady=10)

val_E = Entry(win,font=('Georgia 13'),width=40)
val_E.pack(pady=20)


# def wrapper():
#     write_wb(FILE_PATH, 1, 1, "1")

ttk.Button(win, text="upload", command=lambda: write_wb(FILE_PATH, int(row_num_E.get()), int(col_num_E.get()), str(val_E.get())) ).pack(pady=20)
    

ttk.Button(win, text="write", command=lambda: write_wb(FILE_PATH, int(row_num_E.get()), int(col_num_E.get()), str(val_E.get())) ).pack(pady=20)

win.mainloop()


