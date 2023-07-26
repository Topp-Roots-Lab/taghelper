import mariadb
import sys
import openpyxl
from openpyxl import Workbook

from tkinter import *
from tkinter import ttk, filedialog
from tkinter.filedialog import askopenfile
import os

import getpass

from dbhelper import *

import logging

logging.basicConfig(level=logging.DEBUG)

# def debugLogger(func):
#     def inner(*args, **kwargs):
#         print(f"In {func.__name__}")
#         funcreturn = func(*args,**kwargs)
#         print(f"Exiting {func.__name__}. Returning {funcreturn}.")
#         return funcreturn
#     return inner


# @debugLogger
def getColHeaders(path: str, sheet: str) -> list:
    """
    Return a list of all column headers in a worksheet.

    Parameters:
        str path: The (absolute) path to the Excel workbook being read.
        str sheet: The name of the worksheet tab within the used workbook.

    Return:
        list headers: A list of all values in the first row of a worksheet.
    """

    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[sheet]

    numCols = worksheet.max_column
    headers = []
    for i in range(1, numCols+1):
        headers.append(str(worksheet.cell(row=1, column=i).value))

    return headers



def mapNeededCols(colKey: dict, headers: list):
    """ 
    Creates a mapping of required col names (given by a column key list) to their SPREADSHEET indices.
    Spreadsheet indices start at 1, not 0.

    Parameters:
        dict colKey: A configuration dictionary that describes the required features of a specific database table. Contains requried column names and their respective data types. Pairing is str:class.
        list headers: A list containing all headers in a worksheet.

    Return:
        dict colMap: A dictionary that maps the required column names to their spreadsheet index in the input sheet. Pairing is str:int.
        bool failed: A status boolean that is True if the input sheet is missing a required column and False if all required columns are present.
    """
    colMap = {}
    failed = False
    failedList = []
    for key in colKey.keys():
        try:
            if headers.index(key) >= 0:
                colMap[key] = headers.index(key)+1 #adding 1 to change into spreadsheet index
        except ValueError:
            failed = True
            failedList.append(key)
        
    if failed:
        logging.warning('MISSING FOLLOWING REQUIRED COLUMNS:')
        for col in failedList:
            logging.warning(col)


    print(colMap)
    return colMap, failed

def findUidCol(headers: list):
    failed = False
    colnum = None
    try:
        if headers.index("UID") >= 0:
            colnum = headers.index("UID")+1

    except ValueError:
        logging.warning("No Explicit UID Column Found. Make sure there is an empty column titled 'UID' where UIDs should go.")
        failed = True
    return colnum, failed


def accumDataByRow(colMap: dict, firstRow: int, lastRow: int, path: str, sheet: str):
    """
    Takes a colMap and maps data in required columns to row number.

    Parameters:
        dict colMap: A dictionary that maps the required column names to their spreadsheet index in the input sheet.
        int firstRow: A number that signifies the first row in the input spreadsheet that contains data needing to be uploaded.
        int lastRow: A number that signifies the last row in the input spreadsheet that contains data needing to be uploaded.
        str path: The (explicit) path to the Excel workbook being read.

    Return:
        dict data: A dictionary that maps the row number to a list of data values in that row. Pairing is str:list.
        bool failed: True if some cells were empty. False if all cells contained a value.
    """
    data = {}
    badCells = []
    failed = False
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[sheet]
    for row in range(firstRow, lastRow+1):
        data[str(row)] = []
    for header in colMap:
        for row in range(firstRow, lastRow+1):
            cell = worksheet.cell(row=row, column=colMap[header])
            if cell.value != None:
                data[str(row)] += [cell.value]
            
            else:
                failed = True
                badCells += [(row, header, colMap[header])]
    

    
    if failed:
        logging.warning("The Following cells are empty! Please check sheet and continue after updating!")
        for t in badCells:
            print(f"Row: {t[0]}, Col: {t[2]}, Col Heading: {t[1]}")


    return data, failed




def accumDataByUid(colMap: dict, firstRow: int, lastRow: int, path: str, sheet: str, uidAsData=True):
    """
    Takes a colMap and maps data in required columns to UID.

    Parameters:
        dict colMap: A dictionary that maps the required column names to their spreadsheet index in the input sheet.
        int firstRow: A number that signifies the first row in the input spreadsheet that contains data needing to be uploaded.
        int lastRow: A number that signifies the last row in the input spreadsheet that contains data needing to be uploaded.
        str path: The (explicit) path to the Excel workbook being read.

    Return:
        dict data: A dictionary that maps the UID to a list of data values in that row. Pairing is str:list.
        bool failed: True if some cells were empty. False if all cells contained a value.
    """
    data = {}
    badCells = []
    failed = False
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[sheet]
    for row in range(firstRow, lastRow+1):
        print(str(worksheet.cell(row=row, column=colMap["UID"]).value))
        data[str(worksheet.cell(row=row, column=colMap["UID"]).value)] = []
    for header in colMap:
        if header == "UID" and uidAsData != True:
            continue
        for row in range(firstRow, lastRow+1):
            cell = worksheet.cell(row=row, column=colMap[header])
            if cell.value != None:
                data[str(worksheet.cell(row=row, column=colMap["UID"]).value)] += [cell.value]
            else:
                failed = True
                badCells += [(row, header, colMap[header])]
    

    
    if failed:
        logging.warning("The Following cells are empty! Please check sheet and continue after updating!")
        for t in badCells:
            print(f"Row: {t[0]}, Col: {t[2]}, Col Heading: {t[1]}")


    return data, failed



def validateTypes(data: dict, colKey: dict):
    """
        A function that ensures all values in a data dictionary are the correct type for their respective database column.

        Parameters:
            dict data: A dictionary that maps the rowNum/UID to a list of data values in that row. Pairing is str:list.
            dict colKey: A configuration dictionary that describes the required features of a specific database table. Contains requried column names and their respective data types. Pairing is str:class.

        Return:
            bool failed: True if at least one value is the wrong type. False if all values are the proper type.
    """
    failed = False


    for row in data:
        for i in range(0, len(data[row])):
            if type(data[row][i]) != list(colKey.values())[i]:
                logging.warning(f"Warning: Value {data[row][i]} in Row/UID {row}, Column {list(colKey.keys())[i]} is of improper type. Expected: {list(colKey.values())[i]}. Got {type(data[row][i])}. Please correct this before uploading!")
                failed = True

    return failed


def write_wb(path, firstid, lastid, col_n, sheetName, startrow=1):
    '''
    Writes to a column (col_n) of (path).xlsx the values from (firstid) to (lastid) starting on row (startrow) and incrementing down the column.

    Parameters:
        str path (The path to the .xlsx file to write to. Must end in .xlsx.)
        int firstid (The first uid assigned to the first row to be written to.)
        int lastid (The last id belonging to the last row in this sheet.)
        int col_n (The column to write id values to. Sheet columns are idexed from left to right starting at 1. (A is 1, B is 2, etc.))
        str sheetName (The name of the excel sheet to be witten to.)
        int startrow [optional] (The first row that contains data needing an id value (not col headers etc.). Default value 1 (row 1).)

    Return:
        void

    '''
    wb = openpyxl.load_workbook(path, data_only=True)
    
   
    ws = wb[sheetName]
    rowi = startrow

    assert firstid != None, "Must insert values to generate ids"
    assert lastid != None, "Must insert values to generate ids"

    for i in range(firstid, lastid+1):
        cell = ws.cell(row = rowi, column = col_n)
        cell.value = str(i)
        rowi+=1

    wb.save(path)
    wb.close()
    logging.info("Write successful!")

def mergeDict(dict1: dict, dict2: dict) -> dict:
    return {**dict1, **dict2}

def getFilesOfExt(dir, ext):
   filenames = []
   for file in os.listdir(dir):
       filenames.append(os.path.abspath(file))
    #   if file.endswith(ext) and os.path.isfile(file):
    #      filenames.append(os.path.abspath(file))
    #   else:
    #      continue
   return filenames

